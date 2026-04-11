"""Shared helpers for review export/import workflows."""

from __future__ import annotations

import json
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from finance_tooling.categorization.classify import ClassificationRules, normalize_description

REQUIRED_REVIEW_COLUMNS: tuple[str, ...] = (
    "transaction_id",
    "booking_date",
    "description",
    "amount_native",
    "currency",
    "bank",
    "account_label",
    "category",
    "subcategory",
)
PROJECT_TAGS_COLUMN = "project_tags"
EXISTING_PROJECT_TAGS_COLUMN = "existing_project_tags"
REVIEWED_COLUMN = "reviewed"
REVIEW_COMMENT_COLUMN = "review_comment"
REVIEW_STATUS_COLUMN = "review_status"
REVIEW_GROUP_KEY_COLUMN = "review_group_key"
REVIEW_GROUP_SIZE_COLUMN = "review_group_size"
REVIEW_STATUS_VALUES: tuple[str, ...] = ("todo", "done", "needs_rule", "skip")
NORMALIZED_DESCRIPTION_COLUMN = "normalized_description"
ORIGINAL_CATEGORY_COLUMN = "original_category"
ORIGINAL_SUBCATEGORY_COLUMN = "original_subcategory"
# Backward-compatible alias for older internal references.
FINGERPRINT_COLUMN = NORMALIZED_DESCRIPTION_COLUMN
EDITABLE_REVIEW_COLUMNS = (
    "category",
    "subcategory",
    PROJECT_TAGS_COLUMN,
    REVIEW_STATUS_COLUMN,
    REVIEWED_COLUMN,
    REVIEW_COMMENT_COLUMN,
)


def read_table(path: Path) -> pd.DataFrame:
    """Read CSV/JSON/parquet review or normalized tables."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path)
    if suffix == ".json":
        return pd.read_json(path)
    if suffix == ".xlsx":
        return pd.read_excel(path, engine="openpyxl")
    if suffix == ".parquet":
        return pd.read_parquet(path)
    raise ValueError(
        f"Unsupported table format for {path}; expected .csv, .json, .xlsx, or .parquet"
    )


def _set_column_widths(worksheet) -> None:
    width_overrides = {
        "A": 18,
        "B": 14,
        "C": 54,
        "D": 14,
        "E": 12,
        "F": 12,
        "G": 18,
        "H": 16,
        "I": 20,
        "J": 18,
        "K": 20,
        "L": 18,
        "M": 18,
        "N": 18,
        "O": 18,
        "P": 18,
        "Q": 18,
        "R": 18,
        "S": 12,
        "T": 24,
        "U": 32,
        "V": 16,
        "W": 18,
        "X": 14,
    }
    for column_letter, width in width_overrides.items():
        worksheet.column_dimensions[column_letter].width = width


def _write_instructions_sheet(workbook) -> None:
    worksheet = workbook.create_sheet("instructions")
    worksheet.append(["Finance Tooling Review Workflow"])
    worksheet.append(
        [
            "Edit category/subcategory/project tags as needed, set review_status, "
            "then apply via review-import."
        ]
    )
    worksheet.append([""])
    worksheet.append(["review_status", "meaning"])
    worksheet.append(["todo", "Pending review"])
    worksheet.append(["done", "Reviewed and ready to apply"])
    worksheet.append(["needs_rule", "Pattern should be handled in category rules"])
    worksheet.append(["skip", "Reviewed intentionally without a category change"])
    worksheet.append([""])
    worksheet.append(["Notes"])
    worksheet.append(
        ["- `review_group_key` and `review_group_size` help batch repeated merchants."]
    )
    worksheet.append(
        ["- `reviewed` is a compatibility field derived from review_status on import."]
    )
    worksheet.append(["- Use `review-import --dry-run` before applying changes."])
    worksheet.column_dimensions["A"].width = 28
    worksheet.column_dimensions["B"].width = 96


def taxonomy_label_rows(rules: ClassificationRules) -> list[tuple[str, str | None]]:
    """Return active taxonomy label pairs for workbook validation and helper payloads."""
    label_rows: set[tuple[str, str | None]] = set()
    for raw_entry in rules.taxonomy.values():
        if raw_entry.status != "active":
            continue
        category_label = (raw_entry.category_label or raw_entry.name).strip()
        if not category_label:
            continue
        if raw_entry.subcategory_label is not None:
            label_rows.add((category_label, raw_entry.subcategory_label.strip() or None))
            continue
        if raw_entry.subcategories:
            for subcategory in raw_entry.subcategories:
                normalized_subcategory = str(subcategory).strip()
                label_rows.add((category_label, normalized_subcategory or None))
            continue
        label_rows.add((category_label, None))
    return sorted(label_rows, key=lambda item: (item[0].casefold(), (item[1] or "").casefold()))


def _write_taxonomy_sheet(workbook, *, rules: ClassificationRules | None) -> None:
    worksheet = workbook.create_sheet("taxonomy")
    worksheet.append(["category", "subcategory"])
    rows = [("Uncategorized", None)] if rules is None else taxonomy_label_rows(rules)
    if not rows:
        rows = [("Uncategorized", None)]
    for category, subcategory in rows:
        worksheet.append([category, subcategory or ""])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 24
    worksheet.column_dimensions["B"].width = 28


def _apply_review_validations(worksheet, columns: list[str]) -> None:
    if worksheet.max_row < 2:
        return

    column_indexes = {name: index + 1 for index, name in enumerate(columns)}
    review_status_index = column_indexes.get(REVIEW_STATUS_COLUMN)
    if review_status_index is not None:
        review_status_validation = DataValidation(
            type="list",
            formula1='"todo,done,needs_rule,skip"',
            allow_blank=True,
        )
        review_status_validation.prompt = "Choose review_status"
        review_status_validation.error = "review_status must be todo, done, needs_rule, or skip"
        worksheet.add_data_validation(review_status_validation)
        review_status_validation.add(
            f"{worksheet.cell(row=2, column=review_status_index).coordinate}:"
            f"{worksheet.cell(row=worksheet.max_row, column=review_status_index).coordinate}"
        )

    category_index = column_indexes.get("category")
    if category_index is not None:
        category_validation = DataValidation(
            type="list",
            formula1="=taxonomy!$A$2:$A$2000",
            allow_blank=True,
        )
        category_validation.prompt = "Choose category from taxonomy"
        category_validation.error = "Choose a category listed on the taxonomy sheet"
        worksheet.add_data_validation(category_validation)
        category_validation.add(
            f"{worksheet.cell(row=2, column=category_index).coordinate}:"
            f"{worksheet.cell(row=worksheet.max_row, column=category_index).coordinate}"
        )


def _format_review_workbook(
    path: Path,
    columns: list[str],
    *,
    dark_safe: bool,
    rules: ClassificationRules | None,
) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active
    worksheet.title = "review"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    if dark_safe:
        header_fill = PatternFill(fill_type="solid", fgColor="FF2A3A4E")
        body_fill = PatternFill(fill_type="solid", fgColor="FF1E1E1E")
        editable_fill = PatternFill(fill_type="solid", fgColor="FF5C4B1E")
        header_font = Font(name="Calibri", bold=True, color="FFF5F5F5")
        body_font = Font(name="Calibri", color="FFF5F5F5")
    else:
        header_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
        body_fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
        editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        header_font = Font(name="Calibri", bold=True)
        body_font = Font(name="Calibri")

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    editable_indexes = {
        index + 1
        for index, column_name in enumerate(columns)
        if column_name in EDITABLE_REVIEW_COLUMNS
    }
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for index, cell in enumerate(row, start=1):
            cell.font = body_font
            cell.fill = editable_fill if index in editable_indexes else body_fill

    _set_column_widths(worksheet)
    _apply_review_validations(worksheet, columns)
    _write_instructions_sheet(workbook)
    _write_taxonomy_sheet(workbook, rules=rules)
    workbook.save(path)


def write_table(
    path: Path,
    dataframe: pd.DataFrame,
    *,
    dark_safe: bool = False,
    review_rules: ClassificationRules | None = None,
) -> None:
    """Write CSV/JSON review export tables."""
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        dataframe.to_csv(path, index=False)
        return
    if suffix == ".json":
        path.write_text(
            json.dumps(dataframe.to_dict(orient="records"), indent=2, default=str),
            encoding="utf-8",
        )
        return
    if suffix == ".xlsx":
        dataframe.to_excel(path, index=False, engine="openpyxl")
        _format_review_workbook(
            path,
            dataframe.columns.tolist(),
            dark_safe=dark_safe,
            rules=review_rules,
        )
        return
    raise ValueError(f"Unsupported review output format for {path}; expected .csv, .json, or .xlsx")


def normalize_optional_text(value: object) -> str | None:
    """Normalize optional text-like values from review tables."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return text


def normalize_optional_upper(value: object) -> str | None:
    """Normalize optional text and uppercase it."""
    text = normalize_optional_text(value)
    return text.upper() if text is not None else None


def normalize_reviewed_value(value: object) -> bool:
    """Normalize review marker values from review tables."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y"}


def normalize_review_status_value(value: object, *, reviewed_fallback: object | None = None) -> str:
    """Normalize review status values with backward-compatible reviewed fallback."""
    if isinstance(value, str):
        normalized = value.strip().lower()
        if normalized in REVIEW_STATUS_VALUES:
            return normalized
    return "done" if normalize_reviewed_value(reviewed_fallback) else "todo"


def review_status_is_reviewed(value: object, *, reviewed_fallback: object | None = None) -> bool:
    """Return whether a review status should count as reviewed."""
    return normalize_review_status_value(value, reviewed_fallback=reviewed_fallback) != "todo"


def normalize_optional_decimal(value: object) -> Decimal | None:
    """Normalize optional decimal values from review tables."""
    text = normalize_optional_text(value)
    if text is None:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def normalize_optional_booking_date(value: object) -> str | None:
    """Normalize optional date-like values to ISO format."""
    text = normalize_optional_text(value)
    if text is None:
        return None
    timestamp = pd.to_datetime(text, errors="coerce")
    if pd.isna(timestamp):
        return None
    return timestamp.date().isoformat()


def normalize_project_tags(value: object) -> tuple[str, ...]:
    """Normalize review project tag input into a unique ordered tuple."""
    if value is None:
        return ()
    raw_values: list[str] = []
    if isinstance(value, list | tuple):
        raw_values.extend(str(item).strip() for item in value if str(item).strip())
    else:
        text = str(value).strip()
        if not text or text.lower() == "nan":
            return ()
        raw_values.extend(part.strip() for part in text.split("|"))

    tags: list[str] = []
    seen: set[str] = set()
    for raw in raw_values:
        if not raw:
            continue
        marker = raw.casefold()
        if marker in seen:
            continue
        seen.add(marker)
        tags.append(raw)
    return tuple(tags)


def parse_filter_date(value: str | None, *, label: str) -> date | None:
    """Parse optional inclusive date bounds for review export."""
    if value is None:
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise ValueError(f"Invalid {label}: {value}")
    return parsed.date()


def build_review_group_keys(dataframe: pd.DataFrame) -> pd.Series:
    """Build stable review-group keys for repeated-merchant triage."""
    description_series = dataframe.get(
        "description",
        pd.Series("", index=dataframe.index, dtype="object"),
    ).fillna("")
    normalized_description = description_series.map(lambda value: normalize_description(str(value)))
    bank_series = (
        dataframe.get("bank", pd.Series("", index=dataframe.index, dtype="object"))
        .fillna("")
        .astype(str)
        .str.strip()
        .str.upper()
        .replace("", "UNKNOWN")
    )
    account_series = (
        dataframe.get("account_label", pd.Series("", index=dataframe.index, dtype="object"))
        .fillna("")
        .astype(str)
        .str.strip()
        .replace("", "unlabeled")
    )
    return (
        normalized_description.replace("", "unknown")
        + " | "
        + bank_series
        + " | "
        + account_series
    )
