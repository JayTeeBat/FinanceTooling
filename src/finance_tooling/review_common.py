"""Shared helpers for review export/import workflows."""

from __future__ import annotations

import json
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

REQUIRED_REVIEW_COLUMNS: tuple[str, ...] = (
    "transaction_id",
    "booking_date",
    "description",
    "amount_native",
    "currency",
    "bank",
    "account_label",
    "category",
    "subcategory",
)
PROJECT_TAGS_COLUMN = "project_tags"
EXISTING_PROJECT_TAGS_COLUMN = "existing_project_tags"
REVIEWED_COLUMN = "reviewed"
REVIEW_COMMENT_COLUMN = "review_comment"
FINGERPRINT_COLUMN = "fingerprint"
EDITABLE_REVIEW_COLUMNS = (
    "category",
    "subcategory",
    PROJECT_TAGS_COLUMN,
    REVIEWED_COLUMN,
    REVIEW_COMMENT_COLUMN,
)


def read_table(path: Path) -> pd.DataFrame:
    """Read CSV/JSON/parquet review or normalized tables."""
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return pd.read_csv(path)
    if suffix == ".json":
        return pd.read_json(path)
    if suffix == ".xlsx":
        return pd.read_excel(path, engine="openpyxl")
    if suffix == ".parquet":
        return pd.read_parquet(path)
    raise ValueError(
        f"Unsupported table format for {path}; expected .csv, .json, .xlsx, or .parquet"
    )


def _set_column_widths(worksheet) -> None:
    width_overrides = {
        "A": 18,
        "B": 14,
        "C": 54,
        "D": 14,
        "E": 12,
        "F": 12,
        "G": 18,
        "H": 16,
        "I": 20,
        "J": 18,
        "K": 20,
        "L": 18,
        "M": 18,
        "N": 18,
        "O": 18,
        "P": 18,
        "Q": 18,
        "R": 18,
        "S": 12,
        "T": 24,
        "U": 32,
    }
    for column_letter, width in width_overrides.items():
        worksheet.column_dimensions[column_letter].width = width


def _format_review_workbook(path: Path, columns: list[str], *, dark_safe: bool) -> None:
    workbook = load_workbook(path)
    worksheet = workbook.active
    worksheet.title = "review"
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    if dark_safe:
        header_fill = PatternFill(fill_type="solid", fgColor="FF2A3A4E")
        body_fill = PatternFill(fill_type="solid", fgColor="FF1E1E1E")
        editable_fill = PatternFill(fill_type="solid", fgColor="FF5C4B1E")
        header_font = Font(name="Calibri", bold=True, color="FFF5F5F5")
        body_font = Font(name="Calibri", color="FFF5F5F5")
    else:
        header_fill = PatternFill(fill_type="solid", fgColor="D9E2F3")
        body_fill = PatternFill(fill_type="solid", fgColor="FFFFFFFF")
        editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
        header_font = Font(name="Calibri", bold=True)
        body_font = Font(name="Calibri")

    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    editable_indexes = {
        index + 1
        for index, column_name in enumerate(columns)
        if column_name in EDITABLE_REVIEW_COLUMNS
    }
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for index, cell in enumerate(row, start=1):
            cell.font = body_font
            cell.fill = editable_fill if index in editable_indexes else body_fill

    _set_column_widths(worksheet)
    workbook.save(path)


def write_table(path: Path, dataframe: pd.DataFrame, *, dark_safe: bool = False) -> None:
    """Write CSV/JSON review export tables."""
    path.parent.mkdir(parents=True, exist_ok=True)
    suffix = path.suffix.lower()
    if suffix == ".csv":
        dataframe.to_csv(path, index=False)
        return
    if suffix == ".json":
        path.write_text(
            json.dumps(dataframe.to_dict(orient="records"), indent=2, default=str),
            encoding="utf-8",
        )
        return
    if suffix == ".xlsx":
        dataframe.to_excel(path, index=False, engine="openpyxl")
        _format_review_workbook(path, dataframe.columns.tolist(), dark_safe=dark_safe)
        return
    raise ValueError(f"Unsupported review output format for {path}; expected .csv, .json, or .xlsx")


def normalize_optional_text(value: object) -> str | None:
    """Normalize optional text-like values from review tables."""
    if value is None:
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    return text


def normalize_optional_upper(value: object) -> str | None:
    """Normalize optional text and uppercase it."""
    text = normalize_optional_text(value)
    return text.upper() if text is not None else None


def normalize_reviewed_value(value: object) -> bool:
    """Normalize review marker values from review tables."""
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y"}


def normalize_optional_decimal(value: object) -> Decimal | None:
    """Normalize optional decimal values from review tables."""
    text = normalize_optional_text(value)
    if text is None:
        return None
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def normalize_optional_booking_date(value: object) -> str | None:
    """Normalize optional date-like values to ISO format."""
    text = normalize_optional_text(value)
    if text is None:
        return None
    timestamp = pd.to_datetime(text, errors="coerce")
    if pd.isna(timestamp):
        return None
    return timestamp.date().isoformat()


def normalize_project_tags(value: object) -> tuple[str, ...]:
    """Normalize review project tag input into a unique ordered tuple."""
    if value is None:
        return ()
    raw_values: list[str] = []
    if isinstance(value, list | tuple):
        raw_values.extend(str(item).strip() for item in value if str(item).strip())
    else:
        text = str(value).strip()
        if not text or text.lower() == "nan":
            return ()
        raw_values.extend(part.strip() for part in text.split("|"))

    tags: list[str] = []
    seen: set[str] = set()
    for raw in raw_values:
        if not raw:
            continue
        marker = raw.casefold()
        if marker in seen:
            continue
        seen.add(marker)
        tags.append(raw)
    return tuple(tags)


def parse_filter_date(value: str | None, *, label: str) -> date | None:
    """Parse optional inclusive date bounds for review export."""
    if value is None:
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        raise ValueError(f"Invalid {label}: {value}")
    return parsed.date()
